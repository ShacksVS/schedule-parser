import openpyxl


MAX_ROW_FOR_FACULTY = 15
MAX_ROW = 400
MAX_COL = 6


def load_schedule(file_path):
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        return sheet
    except Exception as e:
        print(f"Error loading the Excel file: {e}")
        return None


def find_faculty(sheet) -> str:
    for row in sheet.iter_rows(min_row=1, max_row=MAX_ROW_FOR_FACULTY, max_col=MAX_COL, values_only=True):
        for cell in row:
            if cell is not None and "Факультет" in cell:
                return cell
    return ""


def find_speciality(sheet) -> str:
    for row in sheet.iter_rows(min_row=1, max_row=MAX_ROW_FOR_FACULTY, max_col=MAX_COL, values_only=True):
        for cell in row:
            if cell is not None and "Спеціальність" in cell:
                return cell
    return ""


def find_row_index(sheet):
    for i, row in enumerate(sheet.iter_rows(min_row=0, max_row=20, max_col=5, values_only=True)):
        if 'День' and 'Час' in row:
            return i + 1
    return -1


def get_clean_data(sheet) -> list:
    temp_time, temp_weekday = None, None

    updated_rows = [find_faculty(sheet), find_speciality(sheet)]

    start_index = find_row_index(sheet)
    if start_index == -1:
        return ["error"]

    for row in sheet.iter_rows(min_row=start_index, max_row=MAX_ROW, max_col=MAX_COL, values_only=True):
        row_list = list(row)

        # Giving correct time for a class without time
        if row_list[2] is not None and row_list[1] is None:
            row_list[1] = temp_time
        elif row_list[1] is not None:
            temp_time = row_list[1]

        # Fill up weekday where None
        if row_list[0] is not None:
            temp_weekday = row_list[0]
        elif temp_weekday is not None:
            row_list[0] = temp_weekday

        updated_row = tuple(row_list)

        if updated_row[1] is not None and updated_row[2]:
            updated_rows.append(updated_row)

    return updated_rows
